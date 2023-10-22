import os
import openpyxl
from datetime import datetime

# Function to check if the Excel file exists and create it if necessary
def create_excel_file():
    if not os.path.exists('library_data.xlsx'):
        workbook = openpyxl.Workbook()
        user_sheet = workbook.active
        user_sheet.title = 'Users'
        user_sheet.append(['Username', 'Password', 'User Type'])
        books_sheet = workbook.create_sheet(title='Books')
        books_sheet.append(['Title', 'Author', 'Availability','user-name'])
        issuance_sheet = workbook.create_sheet(title='Issuance Tracking')
        issuance_sheet.append(['Book Title', 'Borrower', 'Issued Date', 'Returned Date'])
        workbook.save('library_data.xlsx')
        workbook.close()

# Function to register a user (staff or library member)
def register_user(username, password, user_type):
    create_excel_file()
    workbook = openpyxl.load_workbook('library_data.xlsx')
    user_sheet = workbook['Users']

    # Check if the username already exists
    for row in user_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            workbook.close()
            return "Username already exists. Please choose a different one."

    # If the username is unique, add the user to the Excel file
    new_user = [username, password, user_type]
    user_sheet.append(new_user)
    workbook.save('library_data.xlsx')
    workbook.close()
    return "Registration successful."

# Function to log in a user
def login_user(username, password):
    create_excel_file()
    workbook = openpyxl.load_workbook('library_data.xlsx')
    user_sheet = workbook['Users']

    for row in user_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            user_type = row[2]
            workbook.close()
            return f"Login successful. Welcome, {username} ({user_type})."

    workbook.close()
    return "Login failed. Please check your username and password."



# Function to add a book to the catalog
def add_book(title, author):
    create_excel_file()
    workbook = openpyxl.load_workbook('library_data.xlsx')
    books_sheet = workbook['Books']

    # Check if the book already exists
    for row in books_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == title and row[1] == author:
            workbook.close()
            return "Book already exists in the catalog."

    # If the book is not in the catalog, add it
    new_book = [title, author, 'Available']
    books_sheet.append(new_book)
    workbook.save('library_data.xlsx')
    workbook.close()
    return "Book added to the catalog."

# Function to delete a book from the catalog
def delete_book(title, author):
    create_excel_file()
    workbook = openpyxl.load_workbook('library_data.xlsx')
    books_sheet = workbook['Books']

    for row_index, row in enumerate(books_sheet.iter_rows(min_row=2), start=2):
        if row[0].value == title and row[1].value == author:
            books_sheet.delete_rows(row_index)
            workbook.save('library_data.xlsx')
            workbook.close()
            return "Book deleted from the catalog."

    workbook.close()
    return "Book not found in the catalog."

# Function to issue a book to a member
def issue_book(username, title):
    create_excel_file()
    workbook = openpyxl.load_workbook('library_data.xlsx')
    books_sheet = workbook['Books']

    # Check if 'Issuance Tracking' sheet exists, and create it if it doesn't
    if 'Issuance Tracking' not in workbook.sheetnames:
        issuance_sheet = workbook.create_sheet('Issuance Tracking')
        issuance_sheet.append(['Book Title', 'Borrower', 'Issued Date', 'Returned Date'])

    issuance_sheet = workbook['Issuance Tracking']

    issuance_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Find the book by title
    for row_index, row in enumerate(books_sheet.iter_rows(min_row=2), start=2):
        if row[0].value == title:
            availability = row[2].value
            if availability == 'Available':
                # Book is available, issue it to the member
                books_sheet.cell(row=row_index, column=2, value='Not Available')  # Update status to Not Available
                books_sheet.cell(row=row_index, column=4, value=username)  # Store the username of the borrower
                issuance_sheet.append([title, username, issuance_date, ''])  # Record issuance
                workbook.save('library_data.xlsx')
                workbook.close()
                return f"{title} has been issued to {username}."

    # Book not found
    workbook.close()
    return f"Book {title} not found."

# Function to return a book by a member
def return_book(username, title):
    create_excel_file()
    workbook = openpyxl.load_workbook('library_data.xlsx')
    books_sheet = workbook['Books']
    issuance_sheet = workbook['Issuance Tracking']

    return_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Initialize a variable to track the row number
    row_number = 2

    # Create a list to hold the updated book data
    updated_book_data = []

    # Check if the book is in the issuance records and is borrowed by the user
    for row in issuance_sheet.iter_rows(min_row=2, values_only=True):
        book_title, borrower_name, issued_date, returned_date = row
        if book_title == title and borrower_name == username and not returned_date:
            # Book found in issuance records and borrowed by the user
            # Update the return date in the issuance records
            issuance_sheet.cell(row=row_number, column=4, value=return_date)

        # Update the book status to 'Available' and add to the updated_book_data list
        for book_row in books_sheet.iter_rows(min_row=2, values_only=True):
            if book_row[0] == title:
                updated_book_data.append(['Available' if column == book_row[2] else column for column in book_row])
            else:
                updated_book_data.append(list(book_row))

        # Increment the row number
        row_number += 1

    # Clear the content of the "Books" sheet
    for _ in range(2, len(books_sheet['A']) + 1):
        books_sheet.delete_rows(2)

    # Write the updated book data to the "Books" sheet
    for book_data in updated_book_data:
        books_sheet.append(book_data)

    workbook.save('library_data.xlsx')
    workbook.close()
    return f"{title} has been returned by {username}."

    workbook.close()
    return f"{title} is not in your possession or not found in the catalog."



# Main program loop
while True:
    print("Options:")
    print("1. Register a user")
    print("2. Login")
    print("3. Reset password")
    print("4. Add a book")
    print("5. Delete a book")
    print("6. Issue a book (for members)")
    print("7. Return a book (for members)")
    print("8. Search for books")
    print("9. Display book catalog")
    print("10. Quit")
    choice = input("Enter your choice (1/2/3/4/5/6/7/8/9/10): ")

    if choice == '1':
        username = input("Enter a new username: ")
        password = input("Enter a password: ")
        user_type = input("Enter the user type (e.g., Librarian, Member): ")
        result = register_user(username, password, user_type)
        print(result)
    elif choice == '2':
        username = input("Enter your username: ")
        password = input("Enter your password: ")
        result = login_user(username, password)
        print(result)
    elif choice == '3':
        username = input("Enter the username to reset the password: ")
        new_password = input("Enter the new password: ")
        result = reset_password(username, new_password)
        print(result)
    elif choice == '4':
        title = input("Enter the title of the book: ")
        author = input("Enter the author of the book: ")
        result = add_book(title, author)
        print(result)
    elif choice == '5':
        title = input("Enter the title of the book to delete: ")
        author = input("Enter the author of the book to delete: ")
        result = delete_book(title, author)
        print(result)
    elif choice == '6':
        username = input("Enter your username: ")
        title = input("Enter the title of the book to issue: ")
        result = issue_book(username, title)
        print(result)
    elif choice == '7':
        username = input("Enter your username: ")
        title = input("Enter the title of the book to return: ")
        result = return_book(username, title)
        print(result)
    # elif choice == '8':
    #     search_books()
    # elif choice == '9':
    #     display_catalog()
    elif choice == '10':
        break
    else:
        print("Invalid choice. Please choose 1, 2, 3, 4, 5, 6, 7, 8, 9, or 10.")
